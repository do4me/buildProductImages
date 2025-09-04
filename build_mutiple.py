#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
build_Mutiple.py (Python 3.12.3)

What this script does
---------------------
For each row in Products.xlsx (Folder, SKU, ProductName):
1) In that Folder, require: background_mutiple.png, positions_mutiple.txt (5 lines),
   maxheight_mutiple.txt (5 lines), and {SKU}_A.png (current).
2) Read 5 centers from positions_mutiple.txt.
   - Line 3 (index 2) is the center for the current {SKU}_A.png.
   - The other 4 lines are for randomly chosen non-current SKUs' {OTHER}_A.png
     **from the same folder**.
3) Read 5 max heights from maxheight_mutiple.txt.
   If an image exceeds the max height for its slot, scale proportionally to that height.
4) Composite these 5 images onto background_mutiple.png (RGBA).
5) Save the result to ./output/<Folder>/{SKU}_{ProductNameNoSpaces}_mutiple.png,
   preserving the subdirectory structure under ./output/.
   - Non-destructive by default: if the output exists, append " (1)", " (2)", ...
   - Use --overwrite to replace existing output.

CLI
---
python build_Mutiple.py
python build_Mutiple.py --base /path/to/base --xlsx /path/to/Products.xlsx --overwrite --seed 42

Dependencies
------------
- pillow (PIL)
- pandas OR openpyxl (either one to read Excel)

Input file formats
------------------
positions_mutiple.txt: 5 non-empty lines, each a center "x,y" (also accepts "x y", "x:y", "x; y")
maxheight_mutiple.txt: 5 non-empty lines, each an integer (max height in pixels)
"""

from __future__ import annotations

import argparse
import os
import random
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

# Pillow is required for compositing
try:
    from PIL import Image
except ImportError as e:
    raise SystemExit("Pillow (PIL) is required. Please `pip install pillow`.") from e


# ----------------------------- Data structures -----------------------------

@dataclass(frozen=True)
class ProductRow:
    folder: str
    sku: str
    product_name: str  # may be empty


# ----------------------------- Excel loading -------------------------------

def load_products(xlsx_path: Path) -> List[ProductRow]:
    """
    Load Products.xlsx. Requires columns: Folder, SKU, ProductName.
    Trims whitespace; drops rows with missing Folder or SKU.
    """
    required_cols = {"Folder", "SKU", "ProductName"}
    try:
        import pandas as pd  # type: ignore
        df = pd.read_excel(xlsx_path)
        df.columns = [str(c).strip() for c in df.columns]
        missing = required_cols - set(df.columns)
        if missing:
            raise ValueError(f"Missing columns in {xlsx_path.name}: {', '.join(sorted(missing))}")
        df = df.dropna(subset=["Folder", "SKU"]).copy()
        rows: List[ProductRow] = []
        for _, r in df.iterrows():
            folder = str(r["Folder"]).strip()
            sku = str(r["SKU"]).strip()
            pname = "" if ("ProductName" not in r or pd.isna(r["ProductName"])) else str(r["ProductName"]).strip()
            if folder and sku:
                rows.append(ProductRow(folder=folder, sku=sku, product_name=pname))
        return rows
    except ImportError:
        # Fallback: openpyxl
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:
            raise RuntimeError("Install either pandas or openpyxl to read Products.xlsx") from e

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
        missing = required_cols - set(headers)
        if missing:
            raise ValueError(f"Missing columns in {xlsx_path.name}: {', '.join(sorted(missing))}")
        idx = {h: headers.index(h) for h in headers}

        rows: List[ProductRow] = []
        for r in ws.iter_rows(min_row=2):
            folder = r[idx["Folder"]].value if idx["Folder"] < len(r) else None
            sku = r[idx["SKU"]].value if idx["SKU"] < len(r) else None
            pname = r[idx["ProductName"]].value if idx["ProductName"] < len(r) else ""
            if folder is None or sku is None:
                continue
            rows.append(
                ProductRow(
                    folder=str(folder).strip(),
                    sku=str(sku).strip(),
                    product_name="" if pname is None else str(pname).strip(),
                )
            )
        return rows


# ----------------------------- Text parsers --------------------------------

_COORD_PATTERNS: Sequence[re.Pattern[str]] = (
    re.compile(r"^\s*(-?\d+)\s*,\s*(-?\d+)\s*$"),   # "x,y"
    re.compile(r"^\s*(-?\d+)\s+(-?\d+)\s*$"),       # "x y"
    re.compile(r"^\s*(-?\d+)\s*[:;]\s*(-?\d+)\s*$") # "x:y" or "x; y"
)

def parse_xy(line: str) -> Tuple[int, int]:
    """Parse a center coordinate (x, y) from a single line."""
    for pat in _COORD_PATTERNS:
        m = pat.match(line)
        if m:
            return int(m.group(1)), int(m.group(2))
    raise ValueError(f"Invalid coordinate format: '{line}'")

def read_positions(txt_path: Path) -> List[Tuple[int, int]]:
    """Read positions_mutiple.txt; must contain exactly 5 non-empty lines."""
    if not txt_path.exists():
        raise FileNotFoundError(f"Missing {txt_path.name}")
    with txt_path.open("r", encoding="utf-8-sig") as f:
        lines = [ln.strip() for ln in f if ln.strip()]
    if len(lines) != 5:
        raise ValueError(f"{txt_path.name} must contain exactly 5 non-empty lines (got {len(lines)})")
    return [parse_xy(ln) for ln in lines]

def read_maxheights(txt_path: Path) -> List[int]:
    """Read maxheight_mutiple.txt; must contain exactly 5 integer lines."""
    if not txt_path.exists():
        raise FileNotFoundError(f"Missing {txt_path.name}")
    with txt_path.open("r", encoding="utf-8-sig") as f:
        lines = [ln.strip() for ln in f if ln.strip()]
    if len(lines) != 5:
        raise ValueError(f"{txt_path.name} must contain exactly 5 non-empty lines (got {len(lines)})")
    try:
        return [int(v) for v in lines]
    except ValueError as e:
        raise ValueError(f"{txt_path.name} contains non-integer value(s).") from e


# ----------------------------- FS helpers ----------------------------------

def ensure_required_files(folder_path: Path, sku: str) -> Tuple[Path, Path, Path, Path]:
    """
    Ensure presence of required files in the folder:
    - background_mutiple.png
    - positions_mutiple.txt
    - maxheight_mutiple.txt
    - {SKU}_A.png
    """
    bg = folder_path / "background_mutiple.png"
    pos = folder_path / "positions_mutiple.txt"
    maxh = folder_path / "maxheight_mutiple.txt"
    curr = folder_path / f"{sku}_A.png"

    missing = [p.name for p in (bg, pos, maxh, curr) if not p.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required file(s): {', '.join(missing)}")
    return bg, pos, maxh, curr

_SKU_A_RE = re.compile(r"^(?P<sku>.+?)_A\.png$", re.IGNORECASE)

def list_folder_sku_a_images(folder_path: Path) -> Dict[str, Path]:
    """
    Return a mapping {sku: path_to_<sku>_A.png} for all *_A.png files in the folder.
    """
    mapping: Dict[str, Path] = {}
    for p in folder_path.glob("*_A.png"):
        m = _SKU_A_RE.match(p.name)
        if not m:
            continue
        sku = m.group("sku")
        # Use the first occurrence; if duplicates exist, prefer the first seen
        mapping.setdefault(sku, p)
    return mapping

def safe_save_image(img: Image.Image, dst: Path, overwrite: bool = False) -> Path:
    """
    Save RGBA image as PNG to dst. If dst exists and overwrite=False,
    append " (1)", " (2)", ... to the filename.
    """
    dst.parent.mkdir(parents=True, exist_ok=True)
    final_dst = dst
    if final_dst.exists() and not overwrite:
        stem, suffix = final_dst.stem, final_dst.suffix
        i = 1
        while True:
            candidate = final_dst.parent / f"{stem} ({i}){suffix}"
            if not candidate.exists():
                final_dst = candidate
                break
            i += 1
    img.save(final_dst, format="PNG")
    return final_dst


# ----------------------------- Image helpers --------------------------------

def open_rgba(path: Path) -> Image.Image:
    """Open an image and return an RGBA image."""
    img = Image.open(path)
    return img.convert("RGBA")

def paste_with_constraints(
    bg: Image.Image,
    img_path: Path,
    center: Tuple[int, int],
    max_height: int,
) -> None:
    """
    Open `img_path`, resize proportionally if its height exceeds `max_height`,
    and alpha-composite it onto `bg` so that the CENTER of the pasted image
    is located at `center` (x, y).
    """
    if max_height <= 0:
        raise ValueError(f"Invalid max height ({max_height}) for image: {img_path.name}")

    fg = open_rgba(img_path)
    w, h = fg.size
    if h > max_height:
        scale = max_height / h
        new_w = max(1, int(round(w * scale)))
        new_h = max(1, int(round(h * scale)))
        fg = fg.resize((new_w, new_h), Image.LANCZOS)
        w, h = fg.size

    x, y = center
    top_left = (int(round(x - w / 2)), int(round(y - h / 2)))

    if bg.mode != "RGBA":
        raise ValueError("Background must be RGBA for alpha compositing.")
    bg.alpha_composite(fg, dest=top_left)


# ----------------------------- Main routine ---------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Composite current and 4 random non-current *_A.png onto background using per-slot centers and max heights, outputting under ./output/ with preserved subfolder structure."
    )
    parser.add_argument("--base", type=str, default=".", help="Base directory containing all product folders.")
    parser.add_argument("--xlsx", type=str, default="Products.xlsx", help="Path to Products.xlsx.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite outputs if they already exist.")
    parser.add_argument("--seed", type=int, default=None, help="Random seed for reproducible selection of other SKUs.")
    args = parser.parse_args()

    base_dir = Path(args.base).resolve()
    out_base = (base_dir / "output").resolve()
    xlsx_path = (Path(args.xlsx).resolve()
                 if not os.path.isabs(args.xlsx)
                 else Path(args.xlsx))

    if not xlsx_path.exists():
        print(f"[ERROR] Products.xlsx not found: {xlsx_path}", file=sys.stderr)
        return 2

    # PRNG for reproducible selection
    rng = random.Random(args.seed)

    # Load rows
    try:
        rows = load_products(xlsx_path)
    except Exception as e:
        print(f"[ERROR] Failed to read Products.xlsx: {e}", file=sys.stderr)
        return 2

    if not rows:
        print("[INFO] No rows to process.")
        return 0

    ok = skipped = failed = 0

    print(f"[INFO] Base directory : {base_dir}")
    print(f"[INFO] Output base    : {out_base}")
    print(f"[INFO] Total rows     : {len(rows)}\n")

    for idx, row in enumerate(rows, 1):
        folder_name = row.folder
        sku = row.sku
        pname = row.product_name or ""
        pname_no_space = pname.replace(" ", "")

        folder_path = base_dir / folder_name
        print(f"-- Processing {idx}/{len(rows)}: Folder='{folder_name}', SKU='{sku}', ProductName='{pname}'")

        if not folder_path.exists() or not folder_path.is_dir():
            print(f"   [SKIP] Folder not found: {folder_path}")
            skipped += 1
            continue

        # Validate required files for this folder/SKU
        try:
            bg_path, pos_path, maxh_path, curr_img_path = ensure_required_files(folder_path, sku)
        except FileNotFoundError as e:
            print(f"   [SKIP] {e}")
            skipped += 1
            continue

        # Read positions and heights
        try:
            centers = read_positions(pos_path)       # length 5
            max_heights = read_maxheights(maxh_path) # length 5
            print("   [OK] Loaded positions_mutiple.txt and maxheight_mutiple.txt.")
            print(f"       - Slot 3 (index 2): center={centers[2]}, maxH={max_heights[2]} for current {sku}_A.png")
        except Exception as e:
            print(f"   [SKIP] Positions/heights error: {e}")
            skipped += 1
            continue

        # Collect other *_A.png from the SAME folder (exclude current SKU)
        all_a = list_folder_sku_a_images(folder_path)
        if sku in all_a:
            del all_a[sku]
        other_skus = list(all_a.items())  # List[(sku, path)]
        if len(other_skus) < 4:
            print(f"   [SKIP] Not enough non-current *_A.png in this folder (need 4, have {len(other_skus)}).")
            skipped += 1
            continue

        # Randomly pick 4 distinct others
        try:
            selected_others = rng.sample(other_skus, 4)
        except ValueError as e:
            print(f"   [SKIP] Random selection failed: {e}")
            skipped += 1
            continue

        # Open background
        try:
            bg_img = open_rgba(bg_path)
            if bg_img.mode != "RGBA":
                bg_img = bg_img.convert("RGBA")
        except Exception as e:
            print(f"   [FAIL] Cannot open background: {e}")
            failed += 1
            continue

        # Paste current SKU at slot index 2
        try:
            paste_with_constraints(bg_img, curr_img_path, centers[2], max_heights[2])
        except Exception as e:
            print(f"   [FAIL] Failed to paste current '{sku}_A.png': {e}")
            failed += 1
            continue

        # Paste other SKUs at remaining slots [0,1,3,4]
        other_slots = [0, 1, 3, 4]
        try:
            for (slot_idx, (other_sku, other_path)) in zip(other_slots, selected_others):
                if not other_path.exists():
                    raise FileNotFoundError(f"Other SKU image missing: {other_path.name}")
                paste_with_constraints(bg_img, other_path, centers[slot_idx], max_heights[slot_idx])
            print(f"   [OK] Pasted 4 non-current SKUs at slots {other_slots}.")
        except Exception as e:
            print(f"   [FAIL] Failed to paste other SKUs: {e}")
            failed += 1
            continue

        # Save under ./output/<Folder>/
        out_dir = out_base / folder_name
        out_name = f"{sku}_{pname_no_space}_mutiple.png"
        out_path = out_dir / out_name
        try:
            final_path = safe_save_image(bg_img, out_path, overwrite=args.overwrite)
            if final_path == out_path:
                print(f"   [DONE] Saved: {final_path}")
            else:
                print(f"   [DONE] Output existed, saved as: {final_path}")
            ok += 1
        except Exception as e:
            print(f"   [FAIL] Failed to save output: {e}")
            failed += 1
            continue

    # Summary
    print("\n===== Summary =====")
    print(f"Success: {ok}, Skipped: {skipped}, Failed: {failed}")
    return 0 if failed == 0 else 1


# ----------------------------- Entrypoint -----------------------------------

if __name__ == "__main__":
    raise SystemExit(main())
